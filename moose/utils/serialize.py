# -*- coding: utf-8 -*-
import csv
import openpyxl as px

def load_csv(filename):
	content = []
	with open(filename, 'rb') as csvfile:
		csvreader = csv.reader(csvfile)
		for row in csvreader:
			content.append(row)
	return content

def dump_csv(content, filename):
	with open(filename, 'wb') as csvfile:
		csvwriter = csv.writer(csvfile)
		for row in content:
			csvwriter.writerow(row)

# TODO: filters workbook with conditions -
# only return content if its sheet_title, column and row were valid
def load_xlsx(filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.load_workbook(filename=filename)
	workbook = {}
	for ws in wb:
		if sheet_titles and not ws.title in sheet_titles:
			continue
		workbook[ws.title] = []
		for row in ws.rows:
			vals = []
			for cell in row:
				vals.append(cell.value)
			workbook[ws.title].append(vals)
	return workbook


def dump_xlsx(workbook, filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.Workbook()
	actived = False
	if isinstance(workbook, dict):
		for sheet_title, sheet_content in workbook.items():
			if sheet_titles and not sheet_title in sheet_titles:
				continue

			if actived:
				ws = wb.create_sheet(title=sheet_title)
			else:
				ws = wb.active		# have to call active in the first time to create sheet
				ws.title = sheet_title
				actived = True

			for row in sheet_content:
				ws.append(row)
	else:	# syntax sugar for one sheet
		ws = wb.active
		for row in workbook:
			ws.append(row)

	wb.save(filename=filename)
